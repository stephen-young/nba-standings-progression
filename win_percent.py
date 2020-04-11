from openpyxl import load_workbook
from datetime import datetime
from parse import compile
from pandas import DataFrame, date_range, DateOffset, read_excel, to_numeric, concat
from constants import TEAM_COLOURS, PLOT

import pandas
import numpy
import matplotlib.pyplot as plt
pandas.plotting.register_matplotlib_converters()

def get_standings_data_from_spreadsheet(filename):

  workbook = load_workbook(filename,read_only=True,data_only=True)
  sheet = workbook.active
  p = compile('{team} ({win:d}-{loss:d})')
  standings_data = []

  for row in sheet.iter_rows(values_only=True):

    date = datetime.strptime(row[0],"%b %d, %Y").date()
    data = [{'team':'','win':0,'loss':0}]*(len(row)-1)

    for i in range(1,len(row)):
      item = row[i]

      if item is not None:
        r = p.search(item)
        data[i-1] = r.named

    standing_record = {
      "date": date,
      "standing_data": data
    }

    standings_data.append(standing_record)

  return standings_data

def process_spreadsheet_data(data):

  team_list = [item['team'] for item in data[-1]['standing_data']]

  dates = []
  win_percent = {team: [] for team in team_list}

  for item in data:

    dates.append(item['date'])
    standing_data = item['standing_data']

    for team in team_list:

      team_data = [i for i in standing_data if i['team'] == team]

      if len(team_data) == 1:
        team_data = team_data[0]
        win_percent[team].append(team_data['win'] / \
          (team_data['win'] + team_data['loss']))
      else:
        win_percent[team].append(None)

  # package data into a dataframe object

  df = DataFrame(win_percent, index = dates)
  return df

def plot_win_fraction_by_date(data):

  # // put magic numbers into a dict in constants.py

  col_labels = list(data.columns)
  start_date = data.index[0]
  end_date = data.index[-1]

  # // make x-ticks 1 month offsets from start date rather than the start of each month
  date_ticks = date_range(start_date, end_date, freq=DateOffset(months=1))

  fig, axes = plt.subplots(figsize=PLOT['Figure']['Size'])
  axes.set_ylabel('Win fraction (wins / games played to-date)')
  axes.set_ylim(PLOT['Axes']['YMin'], PLOT['Axes']['YMax'])
  axes.set_yticks(PLOT['Axes']['YTick'])

  axes.set_xlabel('Date (YYYY-MM-DD)')
  axes.set_xlim(start_date, end_date)
  axes.set_xticks(date_ticks)

  axes.grid()

  for col in col_labels:
      axes.plot(data.index, data[col], label=col, linewidth=PLOT['Line']['Width'])

  line_array = axes.get_lines()

  # Set line, marker and marker edge colours based on team colours
  for line in line_array:
      team = line.get_label()
      colours = TEAM_COLOURS[team]
      line.set_color(colours['line'])
      line.set_marker('.')
      line.set_markerfacecolor(colours['marker'])
      line.set_markeredgecolor(colours['edge'])
      line.set_markeredgewidth(0.5)
      line.set_markersize(6.0)

  # Set the line style of non-playoff teams to dotted
  for line in line_array[8:]:
      line.set_linestyle(':')

  axes.legend(ncol=PLOT['Legend']['NumCol'])
  fig.set_tight_layout(PLOT['Figure']['TightLayout'])

  return fig
